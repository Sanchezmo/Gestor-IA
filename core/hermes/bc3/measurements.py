"""
BC3 Measurements - Cálculo de mediciones y cantidades.

Proporciona:
- Cálculo de cantidades a partir de geometría
- Agrupación de mediciones por capítulo/subcapítulo
- Cálculo de totales por tipo de recurso
- Generación de estados de mediciones
- Exportación a formatos estándar
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from typing import Any

from .models import (
    BC3Measurement,
    BC3Item,
    BC3Chapter,
    BC3Subchapter,
    BC3Project,
    BC3ResourceType,
    BC3Unit,
    BC3Measurement,
)


# =============================================================================
# MEASUREMENT CALCULATOR
# =============================================================================

class BC3MeasurementCalculator:
    """
    Calculadora de mediciones BC3.
    
    Funciones:
    - Calcular cantidades a partir de geometría básica
    - Agrupar mediciones por capítulo/subcapítulo
    - Calcular totales por tipo de recurso
    - Generar estados de mediciones formateados
    - Exportar a Excel/CSV
    """
    
    def __init__(self, precision: int = 2):
        self.precision = precision
        self._quantize = Decimal(f"0.{'0' * precision}1")  # 0.01, 0.001, etc.
    
    def _quantize_decimal(self, value: Decimal) -> Decimal:
        """Redondear decimal a precisión configurada."""
        return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    
    # =========================================================================
    # GEOMETRY CALCULATIONS
    # =========================================================================
    
    def calculate_area_rectangle(self, width: Decimal, height: Decimal) -> Decimal:
        """Calcular área de rectángulo."""
        return self._quantize_decimal(width * height)
    
    def calculate_area_triangle(self, base: Decimal, height: Decimal) -> Decimal:
        """Calcular área de triángulo."""
        return self._quantize_decimal(base * height / Decimal("2"))
    
    def calculate_area_circle(self, radius: Decimal) -> Decimal:
        """Calcular área de círculo."""
        from decimal import getcontext
        import math
        pi = Decimal(str(math.pi))
        return self._quantize_decimal(pi * radius * radius)
    
    def calculate_volume_prism(self, area: Decimal, height: Decimal) -> Decimal:
        """Calcular volumen de prisma (área x altura)."""
        return self._quantize_decimal(area * height)
    
    def calculate_volume_cylinder(self, radius: Decimal, height: Decimal) -> Decimal:
        """Calcular volumen de cilindro."""
        from decimal import getcontext
        import math
        pi = Decimal(str(math.pi))
        area = pi * radius * radius
        return self._quantize_decimal(area * height)
    
    def calculate_perimeter_rectangle(self, width: Decimal, height: Decimal) -> Decimal:
        """Calcular perímetro de rectángulo."""
        return self._quantize_decimal(Decimal("2") * (width + height))
    
    def calculate_lineal_meters_rectangle(self, width: Decimal, height: Decimal) -> Decimal:
        """Calcular metros lineales de perímetro."""
        return self.calculate_perimeter_rectangle(width, height)
    
    # =========================================================================
    # MEASUREMENT AGGREGATION
    # =========================================================================
    
    def calculate_item_total(self, item: BC3Item) -> Decimal:
        """Calcular precio total de un item (cantidad x precio unitario)."""
        return self._quantize_decimal(item.quantity * item.unit_price)
    
    def calculate_measurement_total(self, measurement: BC3Measurement) -> Decimal:
        """Calcular precio total de una medición."""
        return self._quantize_decimal(measurement.quantity * measurement.unit_price)
    
    def group_measurements_by_chapter(self, measurements: list[BC3Measurement]) -> dict[str, list[BC3Measurement]]:
        """Agrupar mediciones por capítulo."""
        grouped = defaultdict(list)
        for m in measurements:
            grouped[m.chapter].append(m)
        return dict(grouped)
    
    def group_measurements_by_subchapter(self, measurements: list[BC3Measurement]) -> dict[str, list[BC3Measurement]]:
        """Agrupar mediciones por subcapítulo."""
        grouped = defaultdict(list)
        for m in measurements:
            key = f"{m.chapter}.{m.subchapter}" if m.subchapter else m.chapter
            grouped[key].append(m)
        return dict(grouped)
    
    def group_measurements_by_resource_type(self, measurements: list[BC3Measurement]) -> dict[BC3ResourceType, list[BC3Measurement]]:
        """Agrupar mediciones por tipo de recurso."""
        grouped = defaultdict(list)
        for m in measurements:
            grouped[m.resource_type].append(m)
        return dict(grouped)
    
    def calculate_totals_by_resource_type(self, measurements: list[BC3Measurement]) -> dict[BC3ResourceType, dict]:
        """Calcular totales por tipo de recurso."""
        grouped = self.group_measurements_by_resource_type(measurements)
        
        result = {}
        for rt, measurements in grouped.items():
            total_qty = sum(m.quantity for m in measurements)
            total_amount = sum(m.total_price for m in measurements)
            
            result[rt] = {
                "count": len(measurements),
                "total_quantity": sum(m.quantity for m in measurements),
                "total_amount": sum(m.total_price for m in measurements),
                "avg_unit_price": sum(m.unit_price for m in measurements) / len(measurements) if measurements else Decimal("0"),
            }
        
        return result
    
    def calculate_chapter_totals(self, chapter: BC3Chapter) -> dict[str, Decimal]:
        """Calcular totales de un capítulo."""
        total_amount = Decimal("0")
        total_items = 0
        total_measurements = 0
        
        # Items directos
        for item in chapter.items:
            total_items += 1
        
        # Subcapítulos
        for sc in chapter.subchapters:
            total_items += len(sc.items)
            total_measurements += len(sc.measurements)
            total_amount += sc.total_amount
        
        # Mediciones directas del capítulo
        total_measurements += len(chapter.measurements)
        total_amount += sum(m.total_price for m in chapter.measurements)
        
        return {
            "total_items": total_items,
            "total_measurements": total_measurements,
            "total_amount": total_amount,
        }
    
    def calculate_project_totals(self, project: Any) -> dict[str, Any]:  # BC3Project
        """Calcular totales de un proyecto."""
        total_amount = Decimal("0")
        total_chapters = len(project.chapters)
        total_items = 0
        total_measurements = 0
        
        by_resource_type = defaultdict(lambda: {"count": 0, "amount": Decimal("0"), "quantity": Decimal("0")})
        
        for chapter in project.chapters:
            # Items directos
            total_items += len(chapter.items)
            for item in chapter.items:
                by_resource_type[item.resource_type]["count"] += 1
            
            # Subcapítulos
            for sc in chapter.subchapters:
                total_items += len(sc.items)
                for item in sc.items:
                    by_resource_type[item.resource_type]["count"] += 1
                
                # Mediciones
                total_measurements += len(sc.measurements)
                for m in sc.measurements:
                    by_resource_type[m.resource_type]["amount"] += m.total_price
                    by_resource_type[m.resource_type]["quantity"] += m.quantity
            
            # Mediciones directas del capítulo
            total_measurements += len(chapter.measurements)
            for m in chapter.measurements:
                by_resource_type[m.resource_type]["amount"] += m.total_price
                by_resource_type[m.resource_type]["quantity"] += m.quantity
            
            total_amount += sum(c.total_amount for c in project.chapters)
        
        return {
            "total_chapters": total_chapters,
            "total_items": total_items,
            "total_measurements": total_measurements,
            "total_amount": total_amount,
            "by_resource_type": dict(by_resource_type),
        }
    
    # =========================================================================
    # STATE OF MEASUREMENTS (ESTADO DE MEDICIONES)
    # =========================================================================
    
    def generate_state_of_measurements(
        self,
        project: Any,  # BC3Project
        include_breakdown: bool = True,
    ) -> str:
        """
        Generar estado de mediciones formateado.
        
        Formato similar a certificación de obra.
        """
        lines = [
            "=" * 80,
            "ESTADO DE MEDICIONES",
            "=" * 80,
            f"Proyecto: {project.name} ({project.code})",
            f"Fecha: {date.today().strftime('%d/%m/%Y')}",
            "=" * 80,
            "",
        ]
        
        # Resumen global
        totals = self.calculate_project_totals(project)
        lines.extend([
            "RESUMEN GENERAL",
            "-" * 40,
            f"Capítulos: {totals['total_chapters']}",
            f"Items: {totals['total_items']}",
            f"Mediciones: {totals['total_measurements']}",
            f"Importe total: {totals['total_amount']:.2f} €",
            "",
        ])
        
        # Por tipo de recurso
        lines.append("DESGLOSE POR TIPO DE RECURSO")
        lines.append("-" * 40)
        for rt, data in totals["by_resource_type"].items():
            lines.append(f"  {rt.label}: {data['count']} items, {data['amount']:.2f} €")
        lines.append("")
        
        # Detalle por capítulo
        for chapter in project.chapters:
            lines.append(f"CAPÍTULO {chapter.code}: {chapter.description}")
            lines.append("-" * 40)
            
            # Items directos
            for item in chapter.items:
                total = item.quantity * item.unit_price
                lines.append(f"  {item.code} - {item.description}")
                lines.append(f"      {item.quantity} {item.unit.value} x {item.unit_price:.2f} € = {total:.2f} €")
            
            # Subcapítulos
            for sc in chapter.subchapters:
                lines.append(f"  SUBCAPÍTULO {sc.code}: {sc.description}")
                for item in sc.items:
                    total = item.quantity * item.unit_price
                    lines.append(f"    {item.code} - {item.description}")
                    lines.append(f"        {item.quantity} {item.unit.value} x {item.unit_price:.2f} € = {total:.2f} €")
                
                # Mediciones subcapítulo
                for m in sc.measurements:
                    lines.append(f"    MED: {m.item_code} - {m.item_description}")
                    lines.append(f"        {m.quantity} {m.unit.value} x {m.unit_price:.2f} € = {m.total_price:.2f} €")
            
            # Mediciones directas capítulo
            for m in chapter.measurements:
                lines.append(f"  MED: {m.item_code} - {m.item_description}")
                lines.append(f"      {m.quantity} {m.unit.value} x {m.unit_price:.2f} € = {m.total_price:.2f} €")
            
            lines.append("")
        
        # Totales finales
        lines.append("=" * 80)
        lines.append(f"TOTAL PRESUPUESTO: {totals['total_amount']:.2f} €")
        lines.append("=" * 80)
        
        return "\n".join(lines)
    
    def generate_measurement_certificate(
        self,
        project: Any,  # BC3Project
        certificate_number: int,
        period: str,
        previous_certificate_amount: Decimal = Decimal("0"),
    ) -> str:
        """
        Generar certificación de obra (formato certificación parcial).
        """
        lines = [
            "=" * 80,
            f"CERTIFICACIÓN Nº {certificate_number}",
            f"Periodo: {period}",
            "=" * 80,
            "",
        ]
        
        totals = self.calculate_project_totals(project)
        
        lines.extend([
            "RESUMEN ECONÓMICO",
            "-" * 40,
            f"Presupuesto total: {totals['total_amount']:.2f} €",
            f"Certificado anterior: {previous_certificate_amount:.2f} €",
            f"Certificado actual: {(totals['total_amount'] - previous_certificate_amount):.2f} €",
            f"Acumulado: {totals['total_amount']:.2f} €",
            f"Pendiente certificar: {Decimal('0'):.2f} €",  # Simplificado
            "",
        )
        
        # Detalle por capítulo
        for chapter in project.chapters:
            chap_amount = chapter.total_amount
            if chap_amount > 0:
                pct = (chap_amount / totals['total_amount'] * Decimal("100")).quantize(Decimal("0.01")) if totals['total_amount'] > 0 else Decimal("0")
                lines.append(f"Cap. {chapter.code} - {chapter.description}: {chap_amount:.2f} € ({pct}%)")
        
        return "\n".join(lines)
    
    # =========================================================================
    # EXPORT FUNCTIONS
    # =========================================================================
    
    def export_measurements_to_csv(
        self,
        measurements: list[BC3Measurement],
        file_path: str,
        include_breakdown: bool = False,
    ) -> int:
        """Exportar mediciones a CSV."""
        import csv
        
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=";")
            
            # Cabecera
            headers = [
                "Capítulo", "Subcapítulo", "Item", "Descripción",
                "Unidad", "Cantidad", "Precio Unit.", "Importe",
                "Tipo Recurso", "Capítulo", "Subcapítulo"
            ]
            if include_breakdown:
                headers.extend(["Descomposición"])
            
            writer.writerow(headers)
            
            for m in measurements:
                row = [
                    m.chapter,
                    m.subchapter or "",
                    m.item_code,
                    m.item_description,
                    m.unit.value,
                    str(m.quantity),
                    str(m.unit_price),
                    str(m.total_price),
                    m.resource_type.label,
                    m.chapter,
                    m.subchapter or "",
                ]
                if include_breakdown:
                    breakdown_str = "; ".join(
                        f"{b.resource_type.label}: {b.resource_description} ({b.quantity_per_unit} {b.unit.value} x {b.price_per_unit} €)"
                        for b in m.breakdown
                    )
                    row.append(breakdown_str)
                
                writer.writerow(row)
        
        return len(measurements)
    
    def export_project_to_excel(
        self,
        project: Any,  # BC3Project
        file_path: str,
    ) -> int:
        """Exportar proyecto completo a Excel (requiere openpyxl)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl requerido para exportar a Excel. Instale: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        
        # Hoja Resumen
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        
        # Estilos
        header_font = Font(bold=True, size=12)
        header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF", size=11)
        
        # Escribir resumen
        ws_summary.cell(row=1, column=1, value="PROYECTO").font = header_font
        ws_summary.cell(row=1, column=2, value=project.name).font = Font(size=12)
        ws_summary.cell(row=2, column=1, value="CÓDIGO").font = header_font
        ws_summary.cell(row=2, column=2, value=project.code).font = Font(size=12)
        
        totals = self.calculate_project_totals(project)
        
        row = 4
        ws_summary.cell(row=row, column=1, value="TOTAL CAPÍTULOS").font = header_font
        ws_summary.cell(row=row, column=2, value=totals["total_chapters"])
        row += 1
        ws_summary.cell(row=row, column=1, value="TOTAL ITEMS").font = header_font
        ws_summary.cell(row=row, column=2, value=totals["total_items"])
        row += 1
        ws_summary.cell(row=row, column=1, value="TOTAL MEDICIONES").font = header_font
        ws_summary.cell(row=row, column=2, value=totals["total_measurements"])
        row += 1
        ws_summary.cell(row=row, column=1, value="IMPORTE TOTAL").font = header_font
        ws_summary.cell(row=row, column=2, value=float(totals["total_amount"])).font = Font(bold=True, size=12)
        
        # Hoja Mediciones
        ws_measure = wb.create_sheet("Mediciones")
        
        headers = ["Capítulo", "Subcapítulo", "Código", "Descripción", "Unidad", "Cantidad", "P.Unit.", "Importe", "Tipo"]
        for col, header in enumerate(headers, 1):
            cell = ws_measure.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
        
        row = 2
        for chapter in project.chapters:
            for m in chapter.get_all_measurements():
                ws_measure.cell(row=row, column=1, value=chapter.code)
                ws_measure.cell(row=row, column=2, value=m.subchapter or "")
                ws_measure.cell(row=row, column=3, value=m.item_code)
                ws_measure.cell(row=row, column=4, value=m.item_description)
                ws_measure.cell(row=row, column=5, value=m.unit.value)
                ws_measure.cell(row=row, column=6, value=float(m.quantity))
                ws_measure.cell(row=row, column=7, value=float(m.unit_price))
                ws_measure.cell(row=row, column=8, value=float(m.total_price))
                ws_measure.cell(row=row, column=9, value=m.resource_type.label)
                row += 1
        
        # Ajustar anchos
        for ws in [ws_summary, ws_measure]:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        wb.save(file_path)
        return 1


# =============================================================================
# FACTORY FUNCTIONS
# =============================================================================

def create_measurement_calculator(precision: int = 2) -> BC3MeasurementCalculator:
    """Crear calculadora de mediciones."""
    return BC3MeasurementCalculator(precision=precision)


def generate_measurement_state(project: Any, include_breakdown: bool = True) -> str:
    """Generar estado de mediciones formateado."""
    calculator = BC3MeasurementCalculator()
    return calculator.generate_state_of_measurements(project, include_breakdown)


def export_measurements_to_csv(
    measurements: list,
    file_path: str,
    include_breakdown: bool = False,
) -> int:
    """Exportar mediciones a CSV."""
    calculator = BC3MeasurementCalculator()
    return calculator.export_measurements_to_csv(measurements, file_path, include_breakdown)