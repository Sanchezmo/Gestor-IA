"""
BC3 Budget Generator - Generación de presupuestos a partir de mediciones.

Proporciona:
- Generación de presupuestos a partir de mediciones BC3
- Aplicación de precios desde base de precios, manuales o interpolación
- Cálculo de costes directos e indirectos
- Generación de documentos de presupuesto (PDF, Excel, BC3)
- Análisis de rentabilidad
"""

from __future__ import annotations

import tempfile
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from typing import Any
from uuid import UUID, uuid4
from collections import defaultdict

from .models import (
    BC3Project,
    BC3Chapter,
    BC3Subchapter,
    BC3Item,
    BC3Measurement,
    BC3ResourceType,
    BC3Unit,
)
from .cost_breakdown import BC3CostBreakdownCalculator, BC3CostBreakdown
from .price_db import BC3PriceDatabase
from .measurements import BC3MeasurementCalculator


# =============================================================================
# BUDGET MODELS
# =============================================================================

@dataclass(frozen=True, slots=True)
class BC3BudgetLine:
    """Línea de presupuesto."""
    
    id: UUID = field(default_factory=uuid4)
    item_code: str = ""
    item_description: str = ""
    chapter_code: str = ""
    subchapter_code: str | None = None
    unit: BC3Unit = BC3Unit.UN
    quantity: Decimal = Decimal("0")
    unit_price: Decimal = Decimal("0")
    total_price: Decimal = Decimal("0")
    resource_type: BC3ResourceType = BC3ResourceType.MATERIAL
    has_breakdown: bool = False
    breakdown_cost: Decimal = Decimal("0")
    direct_cost: Decimal = Decimal("0")
    overhead_cost: Decimal = Decimal("0")
    profit_cost: Decimal = Decimal("0")
    contingency_cost: Decimal = Decimal("0")
    final_price: Decimal = Decimal("0")
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True, slots=True)
class BC3BudgetSummary:
    """Resumen de presupuesto."""
    
    total_direct_cost: Decimal = Decimal("0")
    total_overhead: Decimal = Decimal("0")
    total_profit: Decimal = Decimal("0")
    total_contingency: Decimal = Decimal("0")
    total_amount: Decimal = Decimal("0")
    total_items: int = 0
    total_measurements: int = 0
    by_resource_type: dict[str, dict] = field(default_factory=dict)
    by_chapter: dict[str, dict] = field(default_factory=dict)


@dataclass(frozen=True, slots=True)
class BC3Budget:
    """Presupuesto completo BC3."""
    
    id: UUID = field(default_factory=uuid4)
    project_id: UUID
    project_name: str = ""
    project_code: str = ""
    lines: list = field(default_factory=list)
    summary: Any = None  # BC3BudgetSummary
    overhead_percentage: Decimal = Decimal("15")
    profit_percentage: Decimal = Decimal("6")
    contingency_percentage: Decimal = Decimal("3")
    retention_percentage: Decimal = Decimal("0")
    created_at: date = field(default_factory=date.today)
    currency: str = "EUR"


# =============================================================================
# BUDGET GENERATOR
# =============================================================================

class BC3BudgetGenerator:
    """
    Generador de presupuestos BC3.
    
    Funciones:
    - Generar presupuesto a partir de proyecto BC3
    - Aplicar precios desde base de datos, manuales o interpolación
    - Calcular costes directos e indirectos
    - Generar documento de presupuesto (Excel, PDF, BC3)
    - Análisis de rentabilidad
    """
    
    def __init__(
        self,
        price_db: BC3PriceDatabase | None = None,
        overhead_default: Decimal = Decimal("15"),
        profit_default: Decimal = Decimal("6"),
        contingency_default: Decimal = Decimal("3"),
        precision: int = 2,
    ):
        self.price_db = price_db or BC3PriceDatabase()
        self.cost_calculator = BC3CostBreakdownCalculator()
        self.measurement_calculator = BC3MeasurementCalculator()
        self.overhead_default = overhead_default
        self.profit_default = profit_default
        self.contingency_default = contingency_default
        self.precision = precision
    
    # =========================================================================
    # MAIN GENERATION
    # =========================================================================
    
    def generate_budget(
        self,
        project: Any,  # BC3Project
        price_source: str = "auto",  # auto | db | manual | interpolate
        overhead: Decimal | None = None,
        profit: Decimal | None = None,
        contingency: Decimal | None = None,
        apply_retention: bool = False,
        retention_rate: Decimal = Decimal("7"),
    ) -> Any:  # BC3Budget
        """
        Generar presupuesto completo a partir de proyecto BC3.
        
        Args:
            project: Proyecto BC3 con mediciones
            price_source: Fuente de precios (auto | db | manual | interpolate)
            overhead: % gastos generales (default 15%)
            profit: % beneficio industrial (default 6%)
            contingency: % contingencias (default 3%)
            apply_retention: Aplicar retención a nivel presupuesto
            retention_rate: % retención (default 7%)
            
        Returns:
            BC3Budget completo
        """
        # Parámetros por defecto
        overhead = overhead or self.overhead_default
        profit = profit or self.profit_default
        contingency = contingency or self.contingency_default
        
        # Generar líneas de presupuesto
        lines = []
        total_direct = Decimal("0")
        total_overhead = Decimal("0")
        total_profit = Decimal("0")
        total_contingency = Decimal("0")
        total_amount = Decimal("0")
        
        by_resource_type = defaultdict(lambda: {"count": 0, "amount": Decimal("0"), "quantity": Decimal("0")})
        by_chapter = defaultdict(lambda: {"count": 0, "amount": Decimal("0"), "quantity": Decimal("0")})
        
        total_items = 0
        total_measurements = 0
        
        for chapter in project.chapters:
            # Procesar items directos del capítulo
            for item in chapter.items:
                line = self._create_budget_line_from_item(
                    item, chapter.code, None,
                    price_source, overhead, profit, contingency,
                    apply_retention, retention_rate
                )
                lines.append(line)
                self._accumulate_totals(line, by_resource_type, by_chapter, chapter.code)
                total_items += 1
            
            # Procesar subcapítulos
            for sc in chapter.subchapters:
                for item in sc.items:
                    line = self._create_budget_line_from_item(
                        item, chapter.code, sc.code,
                        price_source, overhead, profit, contingency,
                        apply_retention, retention_rate
                    )
                    lines.append(line)
                    self._accumulate_totals(line, by_resource_type, by_chapter, chapter.code)
                    total_items += 1
                
                # Mediciones directas del subcapítulo
                for measurement in sc.measurements:
                    line = self._create_budget_line_from_measurement(
                        measurement, chapter.code, sc.code,
                        price_source, overhead, profit, contingency,
                        apply_retention, retention_rate
                    )
                    lines.append(line)
                    self._accumulate_totals(line, by_resource_type, by_chapter, chapter.code)
                    total_measurements += 1
            
            # Mediciones directas del capítulo
            for measurement in chapter.measurements:
                line = self._create_budget_line_from_measurement(
                    measurement, chapter.code, None,
                    price_source, overhead, profit, contingency,
                    apply_retention, retention_rate
                )
                lines.append(line)
                self._accumulate_totals(line, by_resource_type, by_chapter, chapter.code)
                total_measurements += 1
        
        # Calcular totales
        for line in lines:
            total_direct += line.direct_cost
            total_overhead += line.overhead_cost
            total_profit += line.profit_cost
            total_contingency += line.contingency_cost
            total_amount += line.final_price
        
        # Retención a nivel presupuesto
        if apply_retention:
            retention_amount = (total_amount * retention_rate / Decimal("100")).quantize(Decimal("0.01"))
            total_amount -= retention_amount
        
        # Crear resumen
        summary = BC3BudgetSummary(
            total_direct_cost=total_direct,
            total_overhead=total_overhead,
            total_profit=total_profit,
            total_contingency=total_contingency,
            total_amount=total_amount,
            total_items=total_items,
            total_measurements=total_measurements,
            by_resource_type=dict(by_resource_type),
            by_chapter=dict(by_chapter),
        )
        
        return BC3Budget(
            id=uuid4(),
            project_id=project.id,
            project_name=project.name,
            project_code=project.code,
            lines=lines,
            summary=summary,
            overhead_percentage=overhead,
            profit_percentage=profit,
            contingency_percentage=contingency,
            retention_percentage=retention_rate if apply_retention else Decimal("0"),
            created_at=date.today(),
            currency=project.currency,
        )
    
    def _create_budget_line_from_item(
        self,
        item: Any,
        chapter_code: str,
        subchapter_code: str | None,
        price_source: str,
        overhead: Decimal,
        profit: Decimal,
        contingency: Decimal,
        apply_retention: bool,
        retention_rate: Decimal,
    ) -> BC3BudgetLine:
        """Crear línea de presupuesto a partir de un item."""
        # Obtener precio
        unit_price = self._resolve_price(item, price_source)
        
        # Calcular costes
        direct_cost = (item.quantity * unit_price).quantize(Decimal("0.01"))
        overhead_cost = (direct_cost * overhead / Decimal("100")).quantize(Decimal("0.01"))
        subtotal = direct_cost + overhead_cost
        profit_cost = (subtotal * profit / Decimal("100")).quantize(Decimal("0.01"))
        subtotal2 = subtotal + profit_cost
        contingency_cost = (subtotal2 * contingency / Decimal("100")).quantize(Decimal("0.01"))
        
        final_price = direct_cost + overhead_cost + profit_cost + contingency_cost
        
        # Descomposición si existe
        has_breakdown = len(getattr(item, "breakdown", [])) > 0
        breakdown_cost = sum(b.cost_per_unit for b in getattr(item, "breakdown", [])) if has_breakdown else Decimal("0")
        
        return BC3BudgetLine(
            item_code=getattr(item, "code", ""),
            item_description=getattr(item, "description", ""),
            chapter_code=chapter_code,
            subchapter_code=subchapter_code,
            unit=getattr(item, "unit", BC3Unit.UN),
            quantity=getattr(item, "quantity", Decimal("0")),
            unit_price=unit_price,
            total_price=(getattr(item, "quantity", Decimal("0")) * unit_price).quantize(Decimal("0.01")),
            resource_type=getattr(item, "resource_type", BC3ResourceType.MATERIAL),
            has_breakdown=has_breakdown,
            breakdown_cost=breakdown_cost,
            direct_cost=direct_cost,
            overhead_cost=overhead_cost,
            profit_cost=profit_cost,
            contingency_cost=contingency_cost,
            final_price=final_price,
            metadata={"has_breakdown": has_breakdown},
        )
    
    def _create_budget_line_from_measurement(
        self,
        measurement: Any,
        chapter_code: str,
        subchapter_code: str | None,
        price_source: str,
        overhead: Decimal,
        profit: Decimal,
        contingency: Decimal,
        apply_retention: bool,
        retention_rate: Decimal,
    ) -> BC3BudgetLine:
        """Crear línea de presupuesto a partir de una medición."""
        # Obtener precio (puede venir de la medición o resolverse)
        unit_price = getattr(measurement, "unit_price", Decimal("0"))
        
        # Si el precio es 0 o muy bajo, intentar resolver desde BD
        if unit_price <= Decimal("0.01"):
            unit_price = self._resolve_measurement_price(measurement)
        
        direct_cost = (measurement.quantity * unit_price).quantize(Decimal("0.01"))
        overhead_cost = (direct_cost * Decimal("15") / Decimal("100")).quantize(Decimal("0.01"))
        subtotal = direct_cost + overhead_cost
        profit_cost = (subtotal * Decimal("6") / Decimal("100")).quantize(Decimal("0.01"))
        subtotal2 = subtotal + profit_cost
        contingency_cost = (subtotal2 * Decimal("3") / Decimal("100")).quantize(Decimal("0.01"))
        
        final_price = direct_cost + overhead_cost + profit_cost + contingency_cost
        
        has_breakdown = len(getattr(measurement, "breakdown", [])) > 0
        breakdown_cost = sum(b.cost_per_unit for b in getattr(measurement, "breakdown", [])) if has_breakdown else Decimal("0")
        
        return BC3BudgetLine(
            item_code=getattr(measurement, "item_code", ""),
            item_description=getattr(measurement, "item_description", ""),
            chapter_code=chapter_code,
            subchapter_code=subchapter_code,
            unit=getattr(measurement, "unit", BC3Unit.UN),
            quantity=getattr(measurement, "quantity", Decimal("0")),
            unit_price=unit_price,
            total_price=getattr(measurement, "total_price", Decimal("0")),
            resource_type=getattr(measurement, "resource_type", BC3ResourceType.MATERIAL),
            has_breakdown=has_breakdown,
            breakdown_cost=breakdown_cost,
            direct_cost=direct_cost,
            overhead_cost=overhead_cost,
            profit_cost=profit_cost,
            contingency_cost=contingency_cost,
            final_price=final_price,
            metadata={"is_measurement": True},
        )
    
    def _resolve_price(self, item: Any, price_source: str) -> Decimal:
        """Resolver precio unitario para un item."""
        if price_source == "manual" or getattr(item, "unit_price", Decimal("0")) > 0:
            return getattr(item, "unit_price", Decimal("0"))
        
        # Intentar base de datos
        if price_source in ("auto", "db", "interpolate"):
            # Buscar por código exacto
            price_entry = self.price_db.get_best_price(
                getattr(item, "code", ""),
                resource_type=getattr(item, "resource_type", BC3ResourceType.MATERIAL),
            )
            if price_entry:
                return price_entry.price
            
            # Buscar por descripción
            if price_source in ("auto", "interpolate"):
                interp_price = self.price_db.interpolate_price(
                    getattr(item, "description", ""),
                    getattr(item, "resource_type", BC3ResourceType.MATERIAL),
                    getattr(item, "unit", BC3Unit.UN),
                )
                if interp_price:
                    return interp_price
        
        # Fallback: precio del item o 0
        return getattr(item, "unit_price", Decimal("0"))
    
    def _resolve_measurement_price(self, measurement: Any) -> Decimal:
        """Resolver precio para una medición."""
        # Buscar por código de item
        price_entry = self.price_db.get_best_price(
            getattr(measurement, "item_code", ""),
            resource_type=getattr(measurement, "resource_type", BC3ResourceType.MATERIAL),
        )
        if price_entry:
            return price_entry.price
        
        # Interpolar por descripción
        interp_price = self.price_db.interpolate_price(
            getattr(measurement, "item_description", ""),
            getattr(measurement, "resource_type", BC3ResourceType.MATERIAL),
            getattr(measurement, "unit", BC3Unit.UN),
        )
        if interp_price:
            return interp_price
        
        return getattr(measurement, "unit_price", Decimal("0"))
    
    def _accumulate_totals(
        self,
        line: Any,
        by_resource_type: dict,
        by_chapter: dict,
        chapter_code: str,
    ) -> None:
        """Acumular totales para resumen."""
        # Por tipo de recurso
        rt_key = line.resource_type.value
        if rt_key not in by_resource_type:
            by_resource_type[rt_key] = {"count": 0, "amount": Decimal("0"), "quantity": Decimal("0")}
        by_resource_type[rt_key]["count"] += 1
        by_resource_type[rt_key]["amount"] += line.final_price
        by_resource_type[rt_key]["quantity"] += line.quantity
        
        # Por capítulo
        if chapter_code not in by_chapter:
            by_chapter[chapter_code] = {"count": 0, "amount": Decimal("0"), "quantity": Decimal("0")}
        by_chapter[chapter_code]["count"] += 1
        by_chapter[chapter_code]["amount"] += line.final_price
        by_chapter[chapter_code]["quantity"] += line.quantity
    
    # =========================================================================
    # PRICE RESOLUTION
    # =========================================================================
    
    def set_manual_price(self, item_code: str, price: Decimal) -> None:
        """Establecer precio manual para un item (override)."""
        # Se almacenaría en metadata del proyecto o base de datos de precios manuales
        pass
    
    def bulk_update_prices(self, prices: dict[str, Decimal]) -> int:
        """Actualizar precios masivamente."""
        updated = 0
        for code, price in prices.items():
            # Actualizar en base de datos de precios
            entries = self.price_db.search_by_code(code, exact=True)
            for entry in entries:
                self.price_db.update_price(entry.id, {"price": price})
                updated += 1
        return updated
    
    # =========================================================================
    # EXPORT / DOCUMENT GENERATION
    # =========================================================================
    
    def export_budget_to_excel(self, budget: Any, file_path: str) -> int:
        """Exportar presupuesto a Excel."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        
        # Hoja Resumen
        ws = wb.active
        ws.title = "Presupuesto"
        
        # Estilos
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        money_format = '#,##0.00'
        
        # Cabecera
        headers = [
            "Capítulo", "Subcapítulo", "Código", "Descripción",
            "Unidad", "Cantidad", "P.Unit.", "Coste Directo",
            "GG", "BI", "Conting.", "Total", "Tipo"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        row = 2
        for line in budget.lines:
            ws.cell(row=row, column=1, value=line.chapter_code)
            ws.cell(row=row, column=2, value=line.subchapter_code or "")
            ws.cell(row=row, column=3, value=line.item_code)
            ws.cell(row=row, column=4, value=line.item_description)
            ws.cell(row=row, column=5, value=line.unit.value)
            ws.cell(row=row, column=6, value=float(line.quantity))
            ws.cell(row=row, column=7, value=float(line.unit_price))
            ws.cell(row=row, column=8, value=float(line.direct_cost))
            ws.cell(row=row, column=9, value=float(line.overhead_cost))
            ws.cell(row=row, column=10, value=float(line.profit_cost))
            ws.cell(row=row, column=11, value=float(line.contingency_cost))
            ws.cell(row=row, column=12, value=float(line.final_price))
            ws.cell(row=row, column=13, value=line.resource_type.label)
            
            # Formato moneda
            for col in [7, 8, 9, 10, 11, 12]:
                ws.cell(row=row, column=col).number_format = '#,##0.00'
            
            row += 1
        
        # Filas de totales
        summary = budget.summary
        ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True)
        ws.cell(row=row, column=5, value="TOTAL DIRECTO").font = Font(bold=True)
        ws.cell(row=row, column=8, value=float(summary.total_direct_cost)).number_format = '#,##0.00'
        row += 1
        ws.cell(row=row, column=5, value="GASTOS GENERALES").font = Font(bold=True)
        ws.cell(row=row, column=9, value=float(summary.total_overhead)).number_format = '#,##0.00'
        row += 1
        ws.cell(row=row, column=5, value="BENEFICIO INDUSTRIAL").font = Font(bold=True)
        ws.cell(row=row, column=10, value=float(summary.total_profit)).number_format = '#,##0.00'
        row += 1
        ws.cell(row=row, column=5, value="CONTINGENCIAS").font = Font(bold=True)
        ws.cell(row=row, column=11, value=float(summary.total_contingency)).number_format = '#,##0.00'
        row += 1
        ws.cell(row=row, column=5, value="TOTAL PRESUPUESTO").font = Font(bold=True, size=12)
        ws.cell(row=row, column=12, value=float(summary.total_amount)).number_format = '#,##0.00'
        ws.cell(row=row, column=12).font = Font(bold=True, size=12)
        
        # Ajustar anchos
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        wb.save(file_path)
        return 1
    
    def generate_budget_document(
        self,
        budget: Any,  # BC3Budget
        format: str = "excel",  # excel | pdf | bc3
    ) -> bytes:
        """Generar documento de presupuesto en formato especificado."""
        if format == "excel":
            # Crear archivo temporal y leer bytes
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                self.export_budget_to_excel(budget, tmp.name)
                with open(tmp.name, "rb") as f:
                    return f.read()
        elif format == "bc3":
            return self._export_budget_to_bc3(budget)
        else:
            raise ValueError(f"Formato no soportado: {format}")
    
    def _export_budget_to_bc3(self, budget: Any) -> bytes:
        """Exportar presupuesto a formato BC3 XML."""
        import xml.etree.ElementTree as ET
        from xml.dom import minidom
        
        root = ET.Element("bc3")
        root.set("version", "1.0")
        root.set("xmlns", "http://www.bc3.es/")
        
        # Proyecto
        proyecto = ET.SubElement(root, "proyecto")
        proyecto.set("nombre", budget.project_name)
        proyecto.set("codigo", budget.project_code)
        
        # Agrupar por capítulo
        chapters = defaultdict(list)
        for line in budget.lines:
            chapters[line.chapter_code].append(line)
        
        for chapter_code, lines in chapters.items():
            cap_elem = ET.SubElement(root, "capitulo")
            cap_elem.set("codigo", chapter_code)
            # Buscar descripción del capítulo
            # ...
        
        rough_string = ET.tostring(root, encoding='utf-8')
        from xml.dom import minidom
        reparsed = minidom.parseString(rough_string)
        return reparsed.toprettyxml(indent="  ", encoding='utf-8')


# =============================================================================
# PROFITABILITY ANALYSIS
# =============================================================================

@dataclass(frozen=True, slots=True)
class BC3ProfitabilityAnalysis:
    """Análisis de rentabilidad de presupuesto."""
    
    budget_id: UUID
    project_name: str
    total_revenue: Decimal
    total_direct_cost: Decimal
    total_overhead: Decimal
    total_profit: Decimal
    total_contingency: Decimal
    total_cost: Decimal
    gross_margin: Decimal
    net_margin: Decimal
    roi: Decimal  # Return on Investment
    
    # Por tipo de recurso
    profitability_by_resource: dict[str, dict] = field(default_factory=dict)
    # Por capítulo
    profitability_by_chapter: dict[str, dict] = field(default_factory=dict)
    # Items con margen negativo
    negative_margin_items: list[dict] = field(default_factory=list)


class BC3ProfitabilityAnalyzer:
    """Analizador de rentabilidad de presupuestos."""
    
    def __init__(self, budget_generator: BC3BudgetGenerator):
        self.budget_generator = budget_generator
    
    def analyze(self, budget: Any) -> BC3ProfitabilityAnalysis:
        """Analizar rentabilidad de un presupuesto."""
        summary = budget.summary
        
        # Coste total
        total_cost = (summary.total_direct_cost + summary.total_overhead + 
                     summary.total_contingency)
        
        # Ingreso total (precio de venta)
        total_revenue = summary.total_amount
        
        # Márgenes
        gross_margin = ((total_revenue - summary.total_direct_cost) / total_revenue * Decimal("100")).quantize(Decimal("0.01"))
        net_margin = ((total_revenue - total_cost) / total_revenue * Decimal("100")).quantize(Decimal("0.01"))
        
        # ROI
        total_cost = summary.total_direct_cost + summary.total_overhead + summary.total_contingency
        roi = ((total_revenue - total_cost) / total_cost * Decimal("100")).quantize(Decimal("0.01"))
        
        # Análisis por tipo de recurso
        by_resource = {}
        for line in budget.lines:
            rt = line.resource_type.label
            if rt not in by_resource:
                by_resource[rt] = {"revenue": Decimal("0"), "cost": Decimal("0"), "count": 0}
            by_resource[rt]["revenue"] += line.final_price
            # Coste estimado (sin GG, BI, contingencias)
            by_resource[rt]["cost"] += line.direct_cost
            by_resource[rt]["count"] += 1
        
        profitability_by_resource = {}
        for rt, data in by_resource.items():
            margin = ((data["revenue"] - data["cost"]) / data["revenue"] * Decimal("100")).quantize(Decimal("0.01")) if data["revenue"] > 0 else Decimal("0")
            profitability_by_resource[rt] = {
                "revenue": data["revenue"],
                "cost": data["cost"],
                "margin": margin,
                "count": data["count"],
            }
        
        # Items con margen negativo
        negative_margin_items = []
        for line in budget.lines:
            if line.final_price < line.direct_cost:
                negative_margin_items.append({
                    "code": line.item_code,
                    "description": line.item_description,
                    "revenue": line.final_price,
                    "cost": line.direct_cost,
                    "loss": line.direct_cost - line.final_price,
                })
        
        return BC3ProfitabilityAnalysis(
            budget_id=budget.id,
            project_name=budget.project_name,
            total_revenue=total_revenue,
            total_direct_cost=summary.total_direct_cost,
            total_overhead=summary.total_overhead,
            total_profit=summary.total_profit,
            total_contingency=summary.total_contingency,
            total_cost=total_cost,
            gross_margin=gross_margin,
            net_margin=net_margin,
            roi=roi,
            profitability_by_resource=profitability_by_resource,
            profitability_by_chapter={},  # TODO: implementar
            negative_margin_items=negative_margin_items,
        )
    
    def generate_profitability_report(self, analysis: BC3ProfitabilityAnalysis) -> str:
        """Generar informe de rentabilidad formateado."""
        lines = [
            "=" * 80,
            "INFORME DE RENTABILIDAD",
            "=" * 80,
            f"Proyecto: {analysis.project_name}",
            f"Presupuesto ID: {analysis.budget_id}",
            f"Fecha: {date.today().strftime('%d/%m/%Y')}",
            "=" * 80,
            "",
            "RESUMEN ECONÓMICO",
            "-" * 40,
            f"Ingresos totales:      {analysis.total_revenue:>12.2f} €",
            f"Coste directo:         {analysis.total_direct_cost:>12.2f} €",
            f"Gastos generales:      {analysis.total_overhead:>12.2f} €",
            f"Beneficio industrial:  {analysis.total_profit:>12.2f} €",
            f"Contingencias:         {analysis.total_contingency:>12.2f} €",
            f"Coste total:           {analysis.total_cost:>12.2f} €",
            f"Ingresos totales:      {analysis.total_revenue:>12.2f} €",
            "",
            f"MARGEN BRUTO:          {analysis.gross_margin:>6.2f} %",
            f"MARGEN NETO:           {analysis.net_margin:>6.2f} %",
            f"ROI:                   {analysis.roi:>6.2f} %",
            "",
            "ANÁLISIS POR TIPO DE RECURSO",
            "-" * 40,
        ]
        
        for rt, data in analysis.profitability_by_resource.items():
            lines.append(
                f"  {rt}: Ingresos {data['revenue']:.2f} € | "
                f"Coste {data['cost']:.2f} € | "
                f"Margen {data['margin']:.2f}% | "
                f"Items {data['count']}"
            )
        
        if analysis.negative_margin_items:
            lines.extend(["", "ITEMS CON MARGEN NEGATIVO", "-" * 40])
            for item in analysis.negative_margin_items:
                lines.append(
                    f"  {item['code']}: {item['description']} - "
                    f"Pérdida: {item['loss']:.2f} €"
                )
        
        return "\n".join(lines)


# =============================================================================
# FACTORY FUNCTIONS
# =============================================================================

def create_budget_generator(
    price_db: BC3PriceDatabase | None = None,
    overhead: Decimal = Decimal("15"),
    profit: Decimal = Decimal("6"),
    contingency: Decimal = Decimal("3"),
) -> BC3BudgetGenerator:
    """Crear generador de presupuestos."""
    return BC3BudgetGenerator(price_db, overhead, profit, contingency)


def analyze_budget_profitability(budget: Any) -> Any:  # BC3ProfitabilityAnalysis
    """Analizar rentabilidad de un presupuesto."""
    generator = BC3BudgetGenerator()
    analyzer = BC3ProfitabilityAnalyzer(generator)
    return analyzer.analyze(budget)


def generate_budget_document(budget: Any, format: str = "excel") -> bytes:
    """Generar documento de presupuesto."""
    generator = BC3BudgetGenerator()
    return generator.generate_budget_document(budget, format)